import re
from abc import ABCMeta

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from nonebot.log import logger

from ..userdata.userdb import get_all_tables, get_all_rows


def self_adopted_excel_row_width(wb):
    dims = {}

    align = Alignment(horizontal='left', vertical='center')

    # 遍历excel文件的sheet
    for ws in wb.worksheets:
        # 遍历sheet的rows
        for row in ws.rows:
            # 遍历row的cell
            for cell in row:
                if cell.value:
                    cell.alignment = align
                    cell_len = 0.7 * len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                    dims[cell.column] = max(dims.get(cell.column, 0), cell_len)

        # 设置列宽
        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value + 2
    return wb


class BaseSqlExcel(metaclass=ABCMeta):
    def __init__(self):
        self.cursor = None
        self.excel_dict: dict = {}
        self.exception = None
        self.header: list = []
        self.tables = []
        self.sql_constraint = ""

    def dump_excel(self, fn):
        self._create_excel_dict()
        try:
            logger.info('正在绘制excel...')
            wb = self._data_to_workbook()
        except Exception as e:
            self.exception = e
            logger.exception(e)
            logger.warning('绘制excel失败')
            return False
        if wb is None:
            self.exception = NullSqlWarning()
            logger.warning('查询为空')
            return False
        logger.info(f'正在自动调整excel列宽并保存...')
        self_adopted_excel_row_width(wb).save(fn)
        logger.success(f'保存excel完成')
        return True

    def close(self):
        self.cursor.close()

    def _create_excel_dict(self):
        self.tables = get_all_tables(self.cursor)
        for table in self.tables:
            rows = self._get_rows(table)
            # 只有表头,没有数据 则跳过
            if not rows:
                continue
            self.excel_dict.update({
                table: rows
            })

    def _data_to_workbook(self):
        """
        :return: class Workbook
        """
        if self.excel_dict == {}:
            return None
        wb = Workbook()
        header = {i + 1: self.header[i] for i in range(len(self.header))}
        for sheet_name, rows in self.excel_dict.items():
            ws: Worksheet = wb.create_sheet(title=sheet_name)
            # 写入表头
            ws.append(header)
            # 写入数据
            for row in rows:
                ws.append({i + 1: row[i] for i in range(len(row))})
        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
        return wb

    def _get_rows(self, table_name):
        return self.cursor.execute(f"select * from '{table_name}'" + self.sql_constraint + ";").fetchall()

    def __del__(self):
        self.close()


class NullSqlWarning(BaseException):
    def __init__(self):
        super().__init__()
