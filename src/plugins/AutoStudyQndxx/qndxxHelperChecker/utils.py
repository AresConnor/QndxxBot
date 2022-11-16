import urllib.request
from typing import Iterable
from urllib.parse import quote
import pandas as pd
import json
import time
from openpyxl.styles import Alignment, PatternFill
from openpyxl import load_workbook

# =================================================
# 是否启用年级筛选 True False
GRADE_FILTER_SWITCH = True
# 年级: 18 19 20 21
GRADE = 18
# 学院
COLLAGE_NAME = '商学院'
# =================================================

headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.75 Safari/537.36 Edg/100.0.1185.39',
    'origin': 'http://jsutstudy.work',
    'referer': 'http://jsutstudy.work/'
}


def get_class_list(_collage_name: str):
    """

    :param _collage_name: 院名称
    :return: [{'discipline': 专业名,
                  'classes': 此专业下班级的列表
                  }...]
    """
    class_list = []
    url = 'https://dxx.jsutstudy.work/api/getCascadeSelect?college_name=' + quote(str(_collage_name))
    discipline_list = get_data_from_url(url)['data']
    for e in discipline_list:
        url = 'https://dxx.jsutstudy.work/api/getCascadeSelect?discipline_name=' + quote(e) + '&college_name=' + quote(
            _collage_name)
        classes = get_data_from_url(url)['data']

        class_list.append({
            'discipline': e,
            'classes': classes
        })
    return class_list


def get_data_from_url(url: str):
    """

    :param url: 目标url
    :return: 返回json的dict形式
    """
    time.sleep(0.1)
    request = urllib.request.Request(url=url, headers=headers)
    response = urllib.request.urlopen(request)

    content_dict = json.loads(response.read())
    return content_dict


def get_class_data(class_name: str):
    """

    :param class_name: 班级名称 str
    :return: 完成情况 dict
    """
    class_data = get_data_from_url('https://dxx.jsutstudy.work/api/class?class=' + quote(class_name))
    print('获取 ' + class_name + ' 数据')
    return class_data['data']['resource']


def grade_filter(_grade: int, class_list: Iterable, switch):
    """

    :param _grade:
    :param class_list:
    :param switch: 是否开启筛选
    :return:
    """
    filtered_class_list = []
    if switch:
        for disc_dict in class_list:
            for _class in disc_dict['classes']:
                if str(_grade) in _class:
                    filtered_class_list.append(_class)
    if not switch:
        for disc_dict in class_list:
            for _class in disc_dict['classes']:
                filtered_class_list.append(_class)
    print(filtered_class_list)
    return filtered_class_list


def collect_grade_data(class_list:Iterable):
    grade_data = {}
    for _class in class_list:
        grade_data[_class] = get_class_data(_class)
    return grade_data


def get_grade_list():
    _data = get_data_from_url('https://dxx.jsutstudy.work/api/getStatistics?current=1&pageSize=20&')
    tmp = []
    grade_list = []
    for _class in _data['data']['class']:
        tmp.append(_class['name'][:2])
    for i in range(len(tmp)):
        if tmp[i] not in grade_list:
            grade_list.append(tmp[i])
    grade_list.sort()
    return grade_list


def dumps_classes_data():
    print('正在下载数据...')
    data = collect_grade_data(
        grade_filter(_grade=GRADE, class_list=get_class_list(COLLAGE_NAME), switch=GRADE_FILTER_SWITCH))
    print('正在缓存数据...')
    with open('./data.json', 'w', encoding='utf-8') as fp:
        json.dump(data, fp, ensure_ascii=False, indent=4, sort_keys=True)
    fp.close()


def dict2dataframe():
    print('正在转换数据格式...')
    with open('./data.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    total_data = {}
    for _class in data:
        data_t = {
            '年级': [],
            '学号': [],
            '班级': [],
            '姓名': [],
            '专业': [],
            '完成情况': []
        }
        for element in data[_class]:
            data_t['年级'].append(element['grade'])
            data_t['学号'].append(element['id'])
            data_t['班级'].append(_class)
            data_t['姓名'].append(element['name'])
            data_t['专业'].append(element['discipline'])
            data_t['完成情况'].append((element['finish']))
        df_data = pd.DataFrame(data_t.copy())
        df_data.sort_values(by=['学号'], ascending=True, inplace=True)
        total_data[_class] = df_data
    print('正在整合数据...')
    # 合并表 纵向拼接 重排索引
    df_list = []
    for k in total_data:
        df_list.append(total_data[k])
    excel_data = pd.concat(df_list, axis=0, ignore_index=True)
    return excel_data


def dump_excel_file(excel_data: pd.DataFrame, _title, _sheet_name):
    print('正在将数据输出为Excel表格...')
    with pd.ExcelWriter(_title + '.xlsx') as wt:
        excel_data.to_excel(wt, sheet_name=_sheet_name, index=False)


def main():
    dumps_classes_data()
    print('获取表头')
    title = get_data_from_url('https://dxx.jsutstudy.work/api/getStatistics?current=1&pageSize=20&')['data'][
                'lesson_name'] + '青年大学习-' + COLLAGE_NAME + '-完成情况汇总'
    dump_excel_file(dict2dataframe(), _title=title, _sheet_name=COLLAGE_NAME)

    # 设置表单格式
    print('正在格式化Excel...')
    wb = load_workbook(title + '.xlsx')
    for ws_name in wb.sheetnames:
        # 设置单元格格式居中
        for row_cells in wb[ws_name]:
            for cell in row_cells:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # 如果单元格值是未完成，设置单元格背景色为红色
                if cell.value == '未完成':
                    cell.fill = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')

    wb.save(title + '.xlsx')
    wb.close()
    print('完成!')


def get_college_list():
    college_list = []
    data = get_data_from_url('https://dxx.jsutstudy.work/api/getStatistics?current=1&pageSize=20&')
    for e in data['data']['college']:
        if e['grade'] == '总计':
            college_list.append(e['college'])
    return college_list


if __name__ == '__main__':
    print('可以选择一下学院:')
    for e in get_college_list():
        print(e)
    main()

# 获取某个年级青年大学习情况
