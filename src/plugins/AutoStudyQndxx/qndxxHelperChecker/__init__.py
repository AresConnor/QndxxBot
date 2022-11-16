import datetime
import os
import re
from typing import List

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .utils import *
from ..utils import DATA_DIR


def open_excel(fn: str, sheet_name: str) -> Worksheet:
    r = openpyxl.load_workbook(fn)
    return r[sheet_name]


def format_excel(wb):
    dims = {}

    align = Alignment(horizontal='left', vertical='center')

    # 遍历excel文件的sheet
    for ws in wb.worksheets:
        # 遍历sheet的rows
        for row in ws.rows:
            # 遍历row的cell
            for cell in row:
                if cell.value:
                    cell.alignment = align
                    cell_len = 0.7 * len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                    dims[cell.column] = max(dims.get(cell.column, 0), cell_len)
                    if cell.value == '未完成':
                        cell.fill = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')

        # 设置列宽
        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value + 2
    return wb


def helper_checker_run():
    try:
        fn = os.path.join(os.path.split(os.path.realpath(__file__))[0], '青年大学习帮扶列表.xlsx')
        sheet_name = 'Sheet1'
        ws = open_excel(fn, sheet_name)
        _data = [i for i in ws.values]
        header: tuple = _data[0]
        ws_data: List[tuple] = _data[1:]
        classes = {v[header.index('帮扶对象班级')] for v in ws_data if v[header.index('帮扶对象班级')] is not None}

        wb = Workbook()
        ws: Worksheet = wb.active
        # 写入表头
        ws.append({1: '姓名', 2: '帮扶对象', 3: '帮扶对象班级', 4: '帮扶对象学号', 5: '是否帮扶'})

        url_data = collect_grade_data(classes)

        for row in ws_data:
            if None in row:
                break
            for entry in url_data[row[header.index('帮扶对象班级')]]:
                if row[header.index('帮扶对象学号')] == entry['id']:
                    ws.append(
                        {1: row[header.index('姓名')],
                         2: row[header.index('对应帮助的同学')],
                         3: row[header.index('帮扶对象班级')],
                         4: row[header.index('帮扶对象学号')],
                         5: entry['finish']})
                    break
        date_time = datetime.datetime.now()
        file_time = f'{date_time.month}.{date_time.day}'
        fn = file_time + '检查是否帮扶.xlsx'
        format_excel(wb).save(DATA_DIR + fn)

        return fn
    except Exception as exception:
        return exception
